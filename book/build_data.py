# -*- coding: utf-8 -*-
"""把 data.js(xlsx 伪装) 转换为前端数据文件 books.js。
字段缩写成 c=分类 n=书名 a=作者 l=作者等级 s=销量基础，减小体积。
"""
import openpyxl
import json
import shutil
import os

SRC = "data.js"
DST = "books.js"
TMP = "_parse_tmp.xlsx"

shutil.copy(SRC, TMP)
wb = None
try:
    wb = openpyxl.load_workbook(TMP, read_only=True)
    ws = wb[wb.sheetnames[0]]

    books = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] is None or (i == 0 and str(row[0]).strip() == "分类"):
            continue
        books.append({
            "c": str(row[0]).strip(),
            "n": str(row[1]).strip(),
            "a": str(row[2]).strip(),
            "l": str(row[3]).strip(),
            "s": int(row[4]),
        })
finally:
    if wb is not None:
        wb.close()
    if os.path.exists(TMP):
        os.remove(TMP)

# 去重：同名同作者只保留销量基础最高的一条（均订 = s*2/3，随 s 单调）
raw_count = len(books)
best = {}
for b in books:
    key = (b["n"], b["a"])
    cur = best.get(key)
    if cur is None or b["s"] > cur["s"]:
        best[key] = b
books = list(best.values())
print(f"去重：{raw_count} 条 -> {len(books)} 条（同名同作者保留均订最高）")

payload = json.dumps(books, ensure_ascii=False)
with open(DST, "w", encoding="utf-8") as f:
    f.write("/* 书单数据，由 data.js(xlsx) 生成。字段: c=分类 n=书名 a=作者 l=作者等级 s=销量基础 */\n")
    f.write("window.BOOKS = " + payload + ";\n")

print(f"共 {len(books)} 条，写入 {DST}，{len(payload) / 1024:.0f} KB")
